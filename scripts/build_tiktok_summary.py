from __future__ import annotations

import argparse
import gc
import hashlib
import math
import os
import re
import subprocess
import tempfile
import time
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from tqdm import tqdm

try:
    import python_calamine  # noqa: F401

    PANDAS_READ_EXCEL_ENGINE = "calamine"
except ImportError:
    PANDAS_READ_EXCEL_ENGINE = None


ORDER_EXTRA_FEE_FIELDS = [
    "Transaction fee",
    "TikTok Shop commission fee",
    "Seller shipping fee",
    "Affiliate commission",
    "Affiliate Shop Ads commission",
    "Order processing fee",
    "Bonus cashback service fee",
    "Affiliate commission deposit",
    "Shipping Service Fee",
    "Voucher Xtra Service Fee",
    "Pre-order service fee",
    "Affiliate partner commission",
    "LIVE Specials service fee",
    "Campaign resource fee",
    "EAMS Program service fee",
    "GMV Max Coupon",
]

DATAFRAME_CACHE_DIR_NAME = ".cache/build_tiktok_summary"
_PANDAS_DATAFRAME_CACHE: dict[tuple[str, str, str], pd.DataFrame] = {}
tqdm.pandas(ncols=100)

PROGRESS_KWARGS = {"dynamic_ncols": True, "ncols": 100}

ORDER_HEADERS = [
    "店铺",
    "Order ID",
    "Order ID",
    "是否到款",
    "Order Status",
    "Order Substatus",
    "Cancelation/Return Type",
    "Normal or Pre-order",
    "SKU ID",
    "Seller SKU",
    "Product Name",
    "Variation",
    "Quantity",
    "Sku Quantity of return",
    "SKU Unit Original Price",
    "SKU Subtotal Before Discount",
    "SKU Platform Discount",
    "SKU Seller Discount",
    "SKU Subtotal After Discount",
    "Shipping Fee After Discount",
    "Original Shipping Fee",
    "Shipping Fee Seller Discount",
    "Shipping Fee Platform Discount",
    "Payment platform discount",
    "Taxes",
    "Order Amount",
    "Order Refund Amount",
    "Created Time",
    "Paid Time",
    "RTS Time",
    "Shipped Time",
    "Delivered Time",
    "Cancelled Time",
    "Cancel By",
    "Cancel Reason",
    "Fulfillment Type",
    "Warehouse Name",
    "Tracking ID",
    "Delivery Option",
    "Shipping Provider Name",
    "Buyer Message",
    "Buyer Username",
    "Recipient",
    "Phone #",
    "Country",
    "Region",
    "Province",
    "Municipality",
    "Barangay",
    "Detail Address",
    "Additional address information",
    "Payment Method",
    "Weight(kg)",
    "Product Category",
    "Package ID",
    "Seller Note",
    "Checked Status",
    "Checked Marked by",
    "Total revenue",
    "SKU成本(rmb)",
    "SKU总成本(rmb)",
    "Transaction fee",
    "TikTok Shop commission fee",
    "Seller shipping fee",
    "Affiliate commission",
    "Affiliate Shop Ads commission",
    "Order processing fee",
    "Bonus cashback service fee",
    "Affiliate commission deposit",
    "Shipping Service Fee",
    "Voucher Xtra Service Fee",
    "Pre-order service fee",
    "Affiliate partner commission",
    "LIVE Specials service fee",
    "Campaign resource fee",
    "EAMS Program service fee",
    "GMV Max Coupon",
    "广告费",
    "Total fees",
    "结算金额",
    "税费",
    "总计结算金额",
    "净利润(rmb)",
    "备注",
    "取消订单",
    "退款原因",
    "二次销售(0/1)",
]

PAYOUT_HEADERS = [
    "店铺",
    "Order/adjustment ID",
    "Type",
    "Order created time",
    "Order settled time",
    "Currency",
    "Total settlement amount",
    "Total revenue",
    "Subtotal after seller discounts",
    "Subtotal before discounts",
    "Seller discounts",
    "Refund subtotal after seller discounts",
    "Refund subtotal before seller discounts",
    "Refund of seller discounts",
    "Total fees",
    "Transaction fee",
    "TikTok Shop commission fee",
    "Seller shipping fee",
    "Actual shipping fee",
    "Platform shipping fee discount",
    "Customer shipping fee",
    "Refund customer shipping fee",
    "Actual return shipping fee",
    "Shipping fee subsidy",
    "Affiliate commission",
    "Affiliate partner commission",
    "Affiliate Shop Ads commission",
    "Affiliate commission deposit",
    "Affiliate commission refund",
    "Affiliate Partner shop ads commission",
    "GST",
    "SFP service fee",
    "Bonus cashback service fee",
    "LIVE Specials service fee",
    "Voucher Xtra service fee",
    "EAMS Program service fee",
    "Flash Sale service fee",
    "TikTok PayLater program fee",
    "Shipping service fee",
    "Campaign resource fee",
    "Order processing fee",
    "Pre-order service fee",
    "GMV Max Coupon",
    "Ajustment amount",
    "Related order ID",
    "Customer payment",
    "Customer refund",
    "Seller co-funded voucher discount",
    "Refund of seller co-funded voucher discount",
    "Platform discounts",
    "Refund of platform discounts",
    "Platform co-funded voucher discounts",
    "Refund of platform co-funded voucher discounts",
    "Seller shipping fee discount",
    "Estimated package weight (g)",
    "Actual package weight (g)",
]

PAYOUT_FIELD_ALIASES = {
    "Total revenue": "Total revenue",
    "Transaction fee": "Transaction fee",
    "TikTok Shop commission fee": "TikTok Shop commission fee",
    "Seller shipping fee": "Seller shipping fee",
    "Affiliate commission": "Affiliate commission",
    "Affiliate Shop Ads commission": "Affiliate Shop Ads commission",
    "Order processing fee": "Order processing fee",
    "Bonus cashback service fee": "Bonus cashback service fee",
    "Affiliate commission deposit": "Affiliate commission deposit",
    "Shipping Service Fee": "Shipping service fee",
    "Voucher Xtra Service Fee": "Voucher Xtra service fee",
    "Pre-order service fee": "Pre-order service fee",
    "Affiliate partner commission": "Affiliate partner commission",
    "LIVE Specials service fee": "LIVE Specials service fee",
    "Campaign resource fee": "Campaign resource fee",
    "EAMS Program service fee": "EAMS Program service fee",
    "GMV Max Coupon": "GMV Max Coupon",
}

ANALYSIS_HEADERS = [
    "店铺",
    "是否到款",
    "Quantity",
    "Sku Quantity of return",
    "SKU Unit Original Price",
    "SKU Subtotal Before Discount",
    "SKU Platform Discount",
    "SKU Seller Discount",
    "SKU Subtotal After Discount",
    "Shipping Fee After Discount",
    "Original Shipping Fee",
    "Shipping Fee Seller Discount",
    "Shipping Fee Platform Discount",
    "Payment platform discount",
    "Taxes",
    "Order Amount",
    "Order Refund Amount",
    "Cancelled Time",
    "Weight(kg)",
    "Package ID",
    "Seller Note",
    "Total revenue",
    "SKU成本(rmb)",
    "SKU总成本(rmb)",
    "Transaction fee",
    "TikTok Shop commission fee",
    "Seller shipping fee",
    "Affiliate commission",
    "Affiliate Shop Ads commission\t",
    "Order processing fee",
    "Bonus cashback service fee",
    "Affiliate commission deposit",
    "Shipping Service Fee",
    "Voucher Xtra Service Fee",
    "Pre-order service fee",
    "Affiliate partner commission",
    "LIVE Specials service fee",
    "Campaign resource fee",
    "EAMS Program service fee",
    "GMV Max Coupon",
    "广告费",
    "Total fees",
    "结算金额",
    "税费",
    "总计结算金额",
    "总收入(RMB)",
    "净利润(rmb)",
    "利润率",
    "取消订单",
    "Order ID",
    "Order ID.1",
    "SKU ID",
]

AD_IMPORT_SHEET_NAME = "广告费导入表"
AD_IMPORT_GUIDE_SHEET_NAME = "填写说明"
GENERATED_DIR_NAME = "程序生成结果"
AD_IMPORT_HEADERS = [
    "启用",
    "记录类型",
    "月份",
    "店铺",
    "来源单据",
    "来源行号",
    "分摊金额(PHP)",
    "分摊方式",
    "筛选_是否到款",
    "筛选_Order Status",
    "筛选_Seller SKU",
    "筛选_开始时间",
    "筛选_结束时间",
    "调整_Order ID",
    "调整_是否到款",
    "调整_Order Status",
    "调整_Seller SKU",
    "调整_Quantity",
    "Total revenue",
    "Transaction fee",
    "TikTok Shop commission fee",
    "Seller shipping fee",
    "Affiliate commission",
    "Affiliate Shop Ads commission",
    "Order processing fee",
    "Bonus cashback service fee",
    "Affiliate commission deposit",
    "Shipping Service Fee",
    "Voucher Xtra Service Fee",
    "Pre-order service fee",
    "Affiliate partner commission",
    "LIVE Specials service fee",
    "Campaign resource fee",
    "EAMS Program service fee",
    "GMV Max Coupon",
    "广告费",
    "覆盖_Total fees",
    "覆盖_结算金额",
    "覆盖_税费",
    "覆盖_总计结算金额",
    "覆盖_净利润(rmb)",
    "备注",
    "退款原因",
    "二次销售(0/1)",
]
AD_IMPORT_ALLOCATION_TYPES = {"分摊广告费"}
AD_IMPORT_ADJUSTMENT_TYPES = {"新增调整行"}
AD_ALLOCATION_METHODS = {
    "按订单金额": "Order Amount",
    "按Total revenue": "Total revenue",
    "按结算金额": "结算金额",
    "按数量": "Quantity",
    "平均到订单": None,
}

AD_IMPORT_ALWAYS_REQUIRED_HEADERS = {"启用", "记录类型", "月份", "店铺", "来源单据"}
AD_IMPORT_ALLOCATION_REQUIRED_HEADERS = {"分摊金额(PHP)", "分摊方式"}
AD_IMPORT_ALLOCATION_OPTIONAL_HEADERS = {"来源行号", "筛选_是否到款", "筛选_Order Status", "筛选_Seller SKU", "筛选_开始时间", "筛选_结束时间", "备注"}
AD_IMPORT_ADJUSTMENT_REQUIRED_HEADERS = set()
AD_IMPORT_ADJUSTMENT_OPTIONAL_HEADERS = {
    "来源行号",
    "调整_Order ID",
    "调整_是否到款",
    "调整_Order Status",
    "调整_Seller SKU",
    "调整_Quantity",
    "Total revenue",
    "Transaction fee",
    "TikTok Shop commission fee",
    "Seller shipping fee",
    "Affiliate commission",
    "Affiliate Shop Ads commission",
    "Order processing fee",
    "Bonus cashback service fee",
    "Affiliate commission deposit",
    "Shipping Service Fee",
    "Voucher Xtra Service Fee",
    "Pre-order service fee",
    "Affiliate partner commission",
    "LIVE Specials service fee",
    "Campaign resource fee",
    "EAMS Program service fee",
    "GMV Max Coupon",
    "广告费",
    "覆盖_Total fees",
    "覆盖_结算金额",
    "覆盖_税费",
    "覆盖_总计结算金额",
    "覆盖_净利润(rmb)",
    "备注",
    "退款原因",
    "二次销售(0/1)",
}

ORDER_PRIMARY_ID_INDEX = 1
ORDER_SECONDARY_ID_INDEX = 2
ORDER_COLUMN_INDEX = {header: ORDER_HEADERS.index(header) for header in ORDER_HEADERS if header != "Order ID"}
ORDER_COLUMN_INDEX["Order ID"] = ORDER_PRIMARY_ID_INDEX

ANALYSIS_FIRST_VALUE_HEADERS = {"Order ID", "Order ID.1", "SKU ID"}

ANALYSIS_SOURCE_ALIASES = {
    "Affiliate Shop Ads commission\t": ["Affiliate Shop Ads commission\t", "Affiliate Shop Ads commission"],
}

TOP_LEVEL_OWNER_MAP = {
    "品牌美妆": "阳玲",
    "母婴": "阳玲",
    "汽配": "阳玲",
}

PAYOUT_ADJUSTMENT_RULES = {
    "GMV Payment for TikTok Ads": {
        "source_type": "online_ad",
        "note": "广告费",
        "target_header": "广告费",
        "paid_status": "是",
        "match_source_month": True,
    },
    "Withholding tax": {
        "source_type": "withholding_tax",
        "note": "Withholding tax",
        "target_header": "Total revenue",
        "paid_status": "否",
        "match_source_month": True,
    },
    "Platform reimbursement": {
        "source_type": "platform_reimbursement",
        "note": "Platform reimbursement",
        "target_header": "Total revenue",
        "paid_status": "否",
        "match_source_month": True,
    },
    "Logistics reimbursement": {
        "source_type": "logistics_reimbursement",
        "note": "Logistics reimbursement",
        "target_header": "Total revenue",
        "paid_status": "否",
        "match_source_month": True,
    },
    "Violation fee ��settlement fee��": {
        "source_type": "violation_fee",
        "note": "Violation fee ��settlement fee��",
        "target_header": "Total revenue",
        "paid_status": "否",
        "match_source_month": True,
    },
    "Violation fee （settlement fee）": {
        "source_type": "violation_fee",
        "note": "Violation fee （settlement fee）",
        "target_header": "Total revenue",
        "paid_status": "否",
        "match_source_month": True,
    },
}

COLUMN_MISALIGNMENTS = [
    ("Affiliate Shop Ads commission", "Affiliate commission deposit"),
    ("Order processing fee", "Pre-order service fee"),
    ("Bonus cashback service fee", "LIVE Specials service fee"),
    ("Shipping Service Fee", "Campaign resource fee"),
]

MISALIGNMENT_EPSILON = 1e-9


@dataclass(frozen=True)
class SourceFile:
    path: Path
    kind: str
    store_label: str


@dataclass(frozen=True)
class StoreAdjustment:
    store_label: str
    month: int
    order_id: str | None
    amount: float
    note: str
    source_file: str
    source_type: str
    target_header: str = "广告费"
    paid_status: str = "是"


def normalize_header(value: object) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\t", " ").replace("\n", " ")
    text = re.sub(r"\s+", " ", text).strip().lower()
    return text


def get_dataframe_cache_dir() -> Path:
    cache_dir = Path(__file__).resolve().parents[1] / DATAFRAME_CACHE_DIR_NAME
    cache_dir.mkdir(parents=True, exist_ok=True)
    return cache_dir


def log_step(message: str) -> None:
    tqdm.write(message)


def build_dtype_cache_token(dtype: object) -> str:
    if dtype is None:
        return "none"
    if isinstance(dtype, dict):
        return "|".join(f"{key}:{value}" for key, value in sorted(dtype.items(), key=lambda item: str(item[0])))
    return str(dtype)


def build_sheet_cache_path(workbook_path: Path, sheet_name: object, dtype: object) -> Path:
    token = f"{workbook_path.resolve()}|{sheet_name}|{build_dtype_cache_token(dtype)}"
    digest = hashlib.sha1(token.encode("utf-8")).hexdigest()[:16]
    safe_sheet_name = re.sub(r"[^0-9A-Za-z\u4e00-\u9fff_-]+", "_", str(sheet_name))
    return get_dataframe_cache_dir() / f"{workbook_path.stem}.{safe_sheet_name}.{digest}.csv"


def apply_dataframe_dtype(dataframe: pd.DataFrame, dtype: object) -> pd.DataFrame:
    if dtype is None:
        return dataframe
    df = dataframe.copy()
    if isinstance(dtype, dict):
        for column, target_type in dtype.items():
            if column not in df.columns:
                continue
            if target_type is str:
                df[column] = df[column].astype("string")
            else:
                df[column] = df[column].astype(target_type)
        return df
    return df.astype(dtype)


def read_csv_cached(cache_path: Path, dtype: object) -> pd.DataFrame:
    try:
        return pd.read_csv(cache_path, low_memory=False, dtype=dtype)
    except (TypeError, ValueError):
        return apply_dataframe_dtype(pd.read_csv(cache_path, low_memory=False), dtype)


def read_excel_cached(
    workbook_path: Path,
    *,
    sheet_name: str | int = 0,
    dtype: object = None,
    use_csv_cache: bool = True,
) -> pd.DataFrame:
    workbook_path = Path(workbook_path)
    cache_key = (str(workbook_path.resolve()), str(sheet_name), build_dtype_cache_token(dtype))
    cached = _PANDAS_DATAFRAME_CACHE.get(cache_key)
    if cached is not None:
        return cached.copy()

    cache_path = build_sheet_cache_path(workbook_path, sheet_name, dtype) if use_csv_cache else None
    if cache_path is not None and cache_path.exists() and cache_path.stat().st_mtime >= workbook_path.stat().st_mtime:
        df = read_csv_cached(cache_path, dtype)
        _PANDAS_DATAFRAME_CACHE[cache_key] = df
        return df.copy()

    read_kwargs: dict[str, object] = {"sheet_name": sheet_name}
    if dtype is not None:
        read_kwargs["dtype"] = dtype
    if PANDAS_READ_EXCEL_ENGINE is not None:
        read_kwargs["engine"] = PANDAS_READ_EXCEL_ENGINE

    try:
        df = pd.read_excel(workbook_path, **read_kwargs)
    except Exception:
        if "engine" not in read_kwargs:
            raise
        read_kwargs.pop("engine", None)
        df = pd.read_excel(workbook_path, **read_kwargs)

    if cache_path is not None:
        df.to_csv(cache_path, index=False, encoding="utf-8-sig")

    _PANDAS_DATAFRAME_CACHE[cache_key] = df
    return df.copy()


def build_fallback_output_path(output_path: Path) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return output_path.with_name(f"{output_path.stem}_未覆盖_{timestamp}{output_path.suffix}")


def save_workbook_with_fallback(workbook: Workbook, output_path: Path, description: str) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temp_file = tempfile.NamedTemporaryFile(
        prefix=f"{output_path.stem}_",
        suffix=output_path.suffix,
        dir=output_path.parent,
        delete=False,
    )
    temp_file.close()
    temp_path = Path(temp_file.name)
    try:
        workbook.save(temp_path)
        os.replace(temp_path, output_path)
        return output_path
    except PermissionError:
        fallback_path = build_fallback_output_path(output_path)
        if temp_path.exists():
            os.replace(temp_path, fallback_path)
        else:
            workbook.save(fallback_path)
        print(f"警告: {description} 正被其他程序占用，已改存为: {fallback_path}")
        return fallback_path
    finally:
        if temp_path.exists():
            temp_path.unlink(missing_ok=True)


def clean_id(value: object) -> str | None:
    if value in (None, ""):
        return None
    text = str(value).strip()
    if not text:
        return None
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def collect_order_ids(order_files: list[SourceFile], include_canceled: bool) -> set[str]:
    order_ids: set[str] = set()
    for source in tqdm(order_files, desc="Collect order IDs", unit="file", **PROGRESS_KWARGS):
        ws, header_row = open_worksheet(source.path, required_headers=["Order ID", "Order Status"])
        source_index = {normalize_header(header): idx for idx, header in enumerate(header_row)}
        order_id_index = source_index.get(normalize_header("Order ID"))
        status_index = source_index.get(normalize_header("Order Status"))
        if order_id_index is None:
            continue

        for row in ws.iter_rows(min_row=3, values_only=True):
            order_id = clean_id(row[order_id_index] if order_id_index < len(row) else None)
            if not order_id:
                continue
            order_status = row[status_index] if status_index is not None and status_index < len(row) else None
            if not include_canceled and order_status in {"Canceled", "Cancelled"}:
                continue
            order_ids.add(order_id)
    return order_ids


def resolve_payout_order_key(row_values: dict[str, object], valid_order_ids: set[str]) -> str | None:
    related_order_id = clean_id(row_values.get("Related order ID"))
    if related_order_id and related_order_id in valid_order_ids:
        return related_order_id

    adjustment_id = clean_id(row_values.get("Order/adjustment ID"))
    if adjustment_id and adjustment_id in valid_order_ids:
        return adjustment_id

    return None


def to_number(value: object) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def get_datetime(value: object) -> pd.Timestamp | None:
    timestamp = pd.to_datetime(value, errors="coerce")
    if pd.isna(timestamp):
        return None
    return timestamp


def matches_source_month(row_values: dict[str, object], source_month: int) -> bool:
    settled_time = get_datetime(row_values.get("Order settled time"))
    if settled_time is not None:
        return int(settled_time.month) == int(source_month)
    created_time = get_datetime(row_values.get("Order created time"))
    if created_time is not None:
        return int(created_time.month) == int(source_month)
    return False


def numeric_series(df: pd.DataFrame, column: str) -> pd.Series:
    if column not in df.columns:
        return pd.Series(0.0, index=df.index, dtype="float64")
    return pd.to_numeric(df[column], errors="coerce").fillna(0.0)


def non_empty_text(value: object) -> str:
    if value is None:
        return ""
    if pd.isna(value):
        return ""
    return str(value).strip()


def safe_sum(values: Iterable[object]) -> float | None:
    total = 0.0
    has_value = False
    for value in values:
        number = to_number(value)
        if number is None:
            continue
        total += number
        has_value = True
    return total if has_value else None


def clean_cell_value(value: object) -> object:
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub("", value)
    return value


def clean_row(values: Iterable[object]) -> list[object]:
    return [clean_cell_value(value) for value in values]


def first_non_empty(series: pd.Series) -> object:
    for value in series:
        if pd.notna(value) and value != "":
            return clean_cell_value(value)
    return None


def to_excel_value(value: object) -> object:
    if pd.isna(value):
        return None
    return clean_cell_value(value)


def resolve_analysis_source_column(df: pd.DataFrame, header: str) -> str | None:
    candidates = ANALYSIS_SOURCE_ALIASES.get(header, [header])
    for candidate in candidates:
        if candidate in df.columns:
            return candidate
    return None


def build_generated_dir(workspace: Path) -> Path:
    return workspace / "01-TK文件" / GENERATED_DIR_NAME


def autosize_worksheet(worksheet, min_width: int = 12, extra: int = 2) -> None:
    for column_cells in worksheet.columns:
        letter = column_cells[0].column_letter
        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[letter].width = max(min_width, min(max_length + extra, 36))


def style_header_row(worksheet, fill: PatternFill, font: Font, alignment: Alignment, border: Border) -> None:
    for cell in worksheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment
        cell.border = border


def apply_fill_to_headers(worksheet, headers: set[str], fill: PatternFill, font: Font | None = None) -> None:
    target_font = font or Font(bold=True)
    for cell in worksheet[1]:
        if cell.value in headers:
            cell.fill = fill
            cell.font = target_font


def is_truthy(value: object) -> bool:
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return value != 0
    text = str(value).strip().lower()
    return text in {"1", "true", "yes", "y", "是", "启用", "x"}


def set_order_value(row: list[object], header: str, value: object) -> None:
    if header == "Order ID":
        row[ORDER_PRIMARY_ID_INDEX] = value
        row[ORDER_SECONDARY_ID_INDEX] = value
        return
    index = ORDER_COLUMN_INDEX[header]
    row[index] = value


def get_order_value(row: list[object], header: str) -> object:
    if header == "Order ID":
        return row[ORDER_PRIMARY_ID_INDEX]
    return row[ORDER_COLUMN_INDEX[header]]


def append_note(existing: object, addition: object) -> str | None:
    left = "" if existing is None else str(existing).strip()
    right = "" if addition is None else str(addition).strip()
    if not left:
        return right or None
    if not right:
        return left
    if right in left:
        return left
    return f"{left}; {right}"


def distribute_amount(total_amount: float, weights: list[float]) -> list[float]:
    if not weights:
        return []
    positive_weights = [weight if weight > 0 else 0.0 for weight in weights]
    if sum(positive_weights) <= 0:
        positive_weights = [1.0] * len(weights)

    raw_values = [total_amount * weight / sum(positive_weights) for weight in positive_weights]
    rounded = [round(value, 2) for value in raw_values]
    drift = round(total_amount - sum(rounded), 2)
    if rounded:
        rounded[-1] = round(rounded[-1] + drift, 2)
    return rounded


def recompute_order_row(
    row: list[object],
    tax_rate: float,
    exchange_rate: float,
    overrides: dict[str, object] | None = None,
) -> list[object]:
    overrides = overrides or {}
    total_revenue = to_number(get_order_value(row, "Total revenue"))
    fee_values = [to_number(get_order_value(row, field)) for field in ORDER_EXTRA_FEE_FIELDS]
    ad_fee = to_number(get_order_value(row, "广告费"))

    total_fees = to_number(overrides.get("覆盖_Total fees"))
    if total_fees is None:
        total_fees = safe_sum([*fee_values, ad_fee])

    settled_amount = to_number(overrides.get("覆盖_结算金额"))
    if settled_amount is None:
        settled_amount = safe_sum([total_revenue, total_fees])

    tax_fee = to_number(overrides.get("覆盖_税费"))
    if tax_fee is None:
        tax_fee = settled_amount * tax_rate if settled_amount is not None else None

    final_settlement = to_number(overrides.get("覆盖_总计结算金额"))
    if final_settlement is None:
        final_settlement = settled_amount - tax_fee if settled_amount is not None and tax_fee is not None else None

    net_profit = to_number(overrides.get("覆盖_净利润(rmb)"))
    if net_profit is None:
        total_cost = to_number(get_order_value(row, "SKU总成本(rmb)"))
        net_profit = (final_settlement / exchange_rate - total_cost) if final_settlement is not None and total_cost is not None else None

    set_order_value(row, "Total fees", total_fees)
    set_order_value(row, "结算金额", settled_amount)
    set_order_value(row, "税费", tax_fee)
    set_order_value(row, "总计结算金额", final_settlement)
    set_order_value(row, "净利润(rmb)", net_profit)
    return row


def build_empty_adjustment_row(store_label: str, month: object, sequence: int) -> list[object]:
    row = [None] * len(ORDER_HEADERS)
    synthetic_id = f"ADJ-{month or 'NA'}-{sequence:04d}"
    set_order_value(row, "店铺", store_label)
    set_order_value(row, "Order ID", synthetic_id)
    set_order_value(row, "是否到款", "是")
    set_order_value(row, "Order Status", None)
    set_order_value(row, "Quantity", 0)
    set_order_value(row, "SKU成本(rmb)", 0)
    set_order_value(row, "SKU总成本(rmb)", 0)
    set_order_value(row, "广告费", 0)
    set_order_value(row, "二次销售(0/1)", 0)
    return row


def create_store_adjustment_row(
    adjustment: StoreAdjustment,
    sequence: int,
    tax_rate: float,
    exchange_rate: float,
) -> list[object]:
    row = build_empty_adjustment_row(adjustment.store_label, adjustment.month, sequence)
    if adjustment.order_id:
        set_order_value(row, "Order ID", adjustment.order_id)
    set_order_value(row, "是否到款", adjustment.paid_status)
    set_order_value(row, adjustment.target_header, adjustment.amount)
    set_order_value(row, "备注", adjustment.note)
    return clean_row(recompute_order_row(row, tax_rate, exchange_rate))


def build_marketing_store_adjustments(marketing_files: list[SourceFile]) -> list[StoreAdjustment]:
    adjustments: list[StoreAdjustment] = []
    for source in tqdm(marketing_files, desc="Scan marketing files", unit="file", **PROGRESS_KWARGS):
        workbook = load_workbook(source.path, read_only=True, data_only=True)
        month = extract_month(source.path.name) or 2
        for worksheet in workbook.worksheets:
            if worksheet.sheet_state != "visible":
                continue

            preview_rows = list(worksheet.iter_rows(min_row=1, max_row=3, values_only=True))
            if len(preview_rows) < 2:
                continue

            header_row = [str(value).strip() if value is not None else "" for value in preview_rows[1]]
            if "线下广告费金额（PHP）" not in header_row or "平台" not in header_row:
                continue

            source_index = {header: idx for idx, header in enumerate(header_row)}
            for row in worksheet.iter_rows(min_row=3, values_only=True):
                platform = str(row[source_index.get("平台", -1)] or "").strip()
                if platform in {"", "合计："}:
                    continue
                if platform.upper() != "TK":
                    continue

                owner = str(row[source_index.get("业务员", -1)] or "").strip()
                category = str(row[source_index.get("类目", -1)] or "").strip()
                store_label = f"{month}月份-{owner}{category}"

                offline_amount = to_number(row[source_index.get("线下广告费金额（PHP）", -1)]) or 0.0
                if offline_amount:
                    adjustments.append(
                        StoreAdjustment(
                            store_label=store_label,
                            month=month,
                            order_id=None,
                            amount=-offline_amount,
                            note="线下广告费 线下退款",
                            source_file=source.path.name,
                            source_type="offline_ad",
                        )
                    )

                offline_refund_amount = to_number(row[source_index.get("线下退款总金额（PHP）", -1)]) or 0.0
                if offline_refund_amount:
                    adjustments.append(
                        StoreAdjustment(
                            store_label=store_label,
                            month=month,
                            order_id=None,
                            amount=offline_refund_amount,
                            note="线下广告费 线下退款",
                            source_file=source.path.name,
                            source_type="offline_refund",
                        )
                    )
    return adjustments


def apply_system_adjustments(
    order_rows: list[list[object]],
    adjustments: list[StoreAdjustment],
    tax_rate: float,
    exchange_rate: float,
) -> tuple[list[list[object]], dict[str, int]]:
    stats = defaultdict(int)
    sequence = len(order_rows) + 1
    for adjustment in adjustments:
        order_rows.append(create_store_adjustment_row(adjustment, sequence, tax_rate, exchange_rate))
        sequence += 1
        stats[adjustment.source_type] += 1
    return order_rows, dict(stats)


def build_ad_import_workbook(prefill_rows: list[list[object]] | None = None) -> Workbook:
    workbook = Workbook()
    data_sheet = workbook.active
    data_sheet.title = AD_IMPORT_SHEET_NAME
    data_sheet.freeze_panes = "A2"
    data_sheet.append(AD_IMPORT_HEADERS)

    for row in prefill_rows or []:
        data_sheet.append(row)

    header_border = Border(
        left=Side(style="thin", color="00B7B7B7"),
        right=Side(style="thin", color="00B7B7B7"),
        top=Side(style="thin", color="00B7B7B7"),
        bottom=Side(style="thin", color="00B7B7B7"),
    )
    base_header_fill = PatternFill(fill_type="solid", fgColor="00D9E2F3")
    always_required_fill = PatternFill(fill_type="solid", fgColor="00FFD966")
    allocation_required_fill = PatternFill(fill_type="solid", fgColor="00F4B183")
    allocation_optional_fill = PatternFill(fill_type="solid", fgColor="00FFF2CC")
    adjustment_optional_fill = PatternFill(fill_type="solid", fgColor="00DDEBF7")
    override_fill = PatternFill(fill_type="solid", fgColor="00E2F0D9")
    common_font = Font(bold=True, color="001F1F1F")
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    style_header_row(data_sheet, base_header_fill, common_font, center_alignment, header_border)
    apply_fill_to_headers(data_sheet, AD_IMPORT_ALWAYS_REQUIRED_HEADERS, always_required_fill, common_font)
    apply_fill_to_headers(data_sheet, AD_IMPORT_ALLOCATION_REQUIRED_HEADERS, allocation_required_fill, common_font)
    apply_fill_to_headers(data_sheet, AD_IMPORT_ALLOCATION_OPTIONAL_HEADERS, allocation_optional_fill, common_font)
    apply_fill_to_headers(
        data_sheet,
        {header for header in AD_IMPORT_ADJUSTMENT_OPTIONAL_HEADERS if not header.startswith("覆盖_")},
        adjustment_optional_fill,
        common_font,
    )
    apply_fill_to_headers(data_sheet, {header for header in AD_IMPORT_ADJUSTMENT_OPTIONAL_HEADERS if header.startswith("覆盖_")}, override_fill, common_font)

    number_headers = {
        "分摊金额(PHP)",
        "调整_Quantity",
        "Total revenue",
        "Transaction fee",
        "TikTok Shop commission fee",
        "Seller shipping fee",
        "Affiliate commission",
        "Affiliate Shop Ads commission",
        "Order processing fee",
        "Bonus cashback service fee",
        "Affiliate commission deposit",
        "Shipping Service Fee",
        "Voucher Xtra Service Fee",
        "Pre-order service fee",
        "Affiliate partner commission",
        "LIVE Specials service fee",
        "Campaign resource fee",
        "EAMS Program service fee",
        "GMV Max Coupon",
        "广告费",
        "覆盖_Total fees",
        "覆盖_结算金额",
        "覆盖_税费",
        "覆盖_总计结算金额",
        "覆盖_净利润(rmb)",
        "二次销售(0/1)",
    }
    date_headers = {"筛选_开始时间", "筛选_结束时间"}
    for row in data_sheet.iter_rows(min_row=2, max_row=data_sheet.max_row, max_col=data_sheet.max_column):
        for cell in row:
            cell.border = header_border
            cell.alignment = Alignment(vertical="center")
            header = data_sheet.cell(row=1, column=cell.column).value
            if header in number_headers:
                cell.number_format = "#,##0.00"
            elif header in date_headers:
                cell.number_format = "yyyy-mm-dd"

    data_sheet.row_dimensions[1].height = 36
    autosize_worksheet(data_sheet)

    guide_sheet = workbook.create_sheet(AD_IMPORT_GUIDE_SHEET_NAME)
    guide_sheet.freeze_panes = "A2"
    guide_rows = [
        ["字段", "说明"],
        ["颜色图例", "黄色=始终必填；橙色=分摊广告费必填；浅黄=分摊广告费可选筛选；浅蓝=新增调整行可填；浅绿=新增调整行覆盖值"],
        ["记录类型", "只允许：分摊广告费、新增调整行"],
        ["系统默认广告费逻辑", "线上广告费默认来自到款明细表 Type=GMV Payment for TikTok Ads，按结算时间落在当月时生成广告费调整行；线下广告费与线下退款默认来自营销表可见页的线下广告费金额（PHP）/线下退款总金额（PHP），按店铺级调整行写入，不拆分到订单"],
        ["新增系统调整逻辑", "Withholding tax、Platform reimbursement、Logistics reimbursement、Violation fee settlement fee 默认按店铺级 Total revenue 调整行写入；只取结算时间落在文件月份内的记录"],
        ["分摊广告费", "只用于系统未自动识别到的额外广告费场景；按店铺和筛选条件匹配真实订单，再把 分摊金额(PHP) 分摊到 广告费 列"],
        ["新增调整行", "用于补原利润表里状态为空的人工调整行；可直接填写财务列，并允许覆盖 Total fees/结算金额/税费/总计结算金额/净利润"],
        ["分摊方式", "只允许：按订单金额、按Total revenue、按结算金额、按数量、平均到订单"],
        ["金额符号", "费用填负数，补贴/返还填正数，必须与利润表最终方向一致"],
        ["店铺", "必须与程序生成表里的 店铺 完全一致，例如 2月份-刘林长3C"],
        ["覆盖列", "只有新增调整行才需要；留空则脚本按系统公式自动重算"],
        ["最小人工原则", "优先使用系统自动重建结果；只有无法从系统规则推导的补贴/税费/特殊广告费，再填写导入表"],
    ]
    for row in guide_rows:
        guide_sheet.append(row)

    guide_header_fill = PatternFill(fill_type="solid", fgColor="00333F4F")
    guide_header_font = Font(bold=True, color="00FFFFFF")
    style_header_row(guide_sheet, guide_header_fill, guide_header_font, center_alignment, header_border)
    for row in guide_sheet.iter_rows(min_row=2, max_row=guide_sheet.max_row, max_col=guide_sheet.max_column):
        for cell in row:
            cell.border = header_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    autosize_worksheet(guide_sheet, min_width=16, extra=4)
    return workbook


def load_ad_import(ad_import_path: Path | None) -> pd.DataFrame:
    if ad_import_path is None or not ad_import_path.exists():
        return pd.DataFrame(columns=AD_IMPORT_HEADERS)
    df = read_excel_cached(ad_import_path, sheet_name=AD_IMPORT_SHEET_NAME)
    for header in AD_IMPORT_HEADERS:
        if header not in df.columns:
            df[header] = None
    df = df[AD_IMPORT_HEADERS].copy()
    return df[df["启用"].map(is_truthy)].reset_index(drop=True)


def match_filter(row: list[object], rule: pd.Series) -> bool:
    if str(rule.get("店铺") or "").strip() and str(get_order_value(row, "店铺") or "").strip() != str(rule.get("店铺") or "").strip():
        return False

    paid_filter = str(rule.get("筛选_是否到款") or "").strip()
    if paid_filter and str(get_order_value(row, "是否到款") or "").strip() != paid_filter:
        return False

    status_filter = str(rule.get("筛选_Order Status") or "").strip()
    if status_filter and str(get_order_value(row, "Order Status") or "").strip() != status_filter:
        return False

    seller_sku_filter = clean_id(rule.get("筛选_Seller SKU"))
    if seller_sku_filter and clean_id(get_order_value(row, "Seller SKU")) != seller_sku_filter:
        return False

    start_time = pd.to_datetime(rule.get("筛选_开始时间"), errors="coerce")
    if pd.notna(start_time):
        created_time = pd.to_datetime(get_order_value(row, "Created Time"), errors="coerce")
        if pd.isna(created_time) or created_time < start_time:
            return False

    end_time = pd.to_datetime(rule.get("筛选_结束时间"), errors="coerce")
    if pd.notna(end_time):
        created_time = pd.to_datetime(get_order_value(row, "Created Time"), errors="coerce")
        if pd.isna(created_time) or created_time > end_time:
            return False

    return True


def allocation_weight(row: list[object], method: str) -> float:
    source_header = AD_ALLOCATION_METHODS.get(method)
    if source_header is None:
        return 1.0
    value = to_number(get_order_value(row, source_header))
    if value is None:
        return 0.0
    return max(value, 0.0)


def apply_ad_import(
    order_rows: list[list[object]],
    ad_import_df: pd.DataFrame,
    tax_rate: float,
    exchange_rate: float,
) -> tuple[list[list[object]], dict[str, int]]:
    if ad_import_df.empty:
        return order_rows, {"allocation_rules": 0, "adjustment_rows": 0}

    allocation_rule_count = 0
    adjustment_row_count = 0

    rules = ad_import_df.to_dict(orient="records")
    for rule in tqdm(rules, desc="Apply ad import rules", unit="rule", **PROGRESS_KWARGS):
        record_type = str(rule.get("记录类型") or "").strip()
        if record_type in AD_IMPORT_ALLOCATION_TYPES:
            method = str(rule.get("分摊方式") or "平均到订单").strip() or "平均到订单"
            if method not in AD_ALLOCATION_METHODS:
                raise ValueError(f"不支持的广告费分摊方式: {method}")

            selected_rows = [row for row in order_rows if match_filter(row, rule)]
            if not selected_rows:
                raise ValueError(f"广告费分摊规则未匹配到任何订单: 店铺={rule.get('店铺')} 来源单据={rule.get('来源单据')}")

            amount = to_number(rule.get("分摊金额(PHP)"))
            if amount is None:
                raise ValueError("广告费分摊规则缺少 分摊金额(PHP)")

            weights = [allocation_weight(row, method) for row in selected_rows]
            distributed = distribute_amount(amount, weights)
            allocation_note = append_note(rule.get("备注"), rule.get("来源单据"))

            for row, share in zip(selected_rows, distributed):
                current_ad_fee = to_number(get_order_value(row, "广告费")) or 0.0
                set_order_value(row, "广告费", round(current_ad_fee + share, 2))
                if allocation_note:
                    set_order_value(row, "备注", append_note(get_order_value(row, "备注"), allocation_note))
                recompute_order_row(row, tax_rate, exchange_rate)

            allocation_rule_count += 1
            continue

        if record_type in AD_IMPORT_ADJUSTMENT_TYPES:
            adjustment_row = build_empty_adjustment_row(rule.get("店铺"), rule.get("月份"), len(order_rows) + adjustment_row_count + 1)
            if str(rule.get("调整_Order ID") or "").strip():
                set_order_value(adjustment_row, "Order ID", str(rule.get("调整_Order ID")).strip())
            if str(rule.get("调整_是否到款") or "").strip():
                set_order_value(adjustment_row, "是否到款", str(rule.get("调整_是否到款")).strip())
            if pd.notna(rule.get("调整_Order Status")):
                set_order_value(adjustment_row, "Order Status", rule.get("调整_Order Status"))
            if pd.notna(rule.get("调整_Seller SKU")):
                set_order_value(adjustment_row, "Seller SKU", rule.get("调整_Seller SKU"))
            if pd.notna(rule.get("调整_Quantity")):
                set_order_value(adjustment_row, "Quantity", to_number(rule.get("调整_Quantity")) or 0)

            for header in ["Total revenue", *ORDER_EXTRA_FEE_FIELDS, "广告费", "备注", "退款原因", "二次销售(0/1)"]:
                if pd.notna(rule.get(header)) and rule.get(header) != "":
                    set_order_value(adjustment_row, header, rule.get(header))

            adjustment_row = recompute_order_row(
                adjustment_row,
                tax_rate,
                exchange_rate,
                overrides={
                    "覆盖_Total fees": rule.get("覆盖_Total fees"),
                    "覆盖_结算金额": rule.get("覆盖_结算金额"),
                    "覆盖_税费": rule.get("覆盖_税费"),
                    "覆盖_总计结算金额": rule.get("覆盖_总计结算金额"),
                    "覆盖_净利润(rmb)": rule.get("覆盖_净利润(rmb)"),
                },
            )
            order_rows.append(clean_row(adjustment_row))
            adjustment_row_count += 1
            continue

        raise ValueError(f"不支持的记录类型: {record_type}")

    return order_rows, {
        "allocation_rules": allocation_rule_count,
        "adjustment_rows": adjustment_row_count,
    }


def export_ad_import_template(output_path: Path) -> None:
    workbook = build_ad_import_workbook([[
        0,
        "分摊广告费",
        2,
        "2月份-刘林长3C",
        "刘林长 菲律宾 2月营销费+退款统计.xlsx",
        3,
        -3549.42,
        "按Total revenue",
        "是",
        "Completed",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        "2月广告费自动分摊",
        None,
        None,
    ]])
    workbook[AD_IMPORT_SHEET_NAME].append([
        0,
        "新增调整行",
        2,
        "2月份-刘林长3C",
        "人工补录",
        1,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        "ADJ-SAMPLE-0001",
        "是",
        None,
        None,
        0,
        0,
        0,
        0,
        0,
        0,
        0,
        0,
        0,
        0,
        0,
        0,
        0,
        0,
        0,
        0,
        0,
        0,
        0,
        None,
        None,
        -127.71,
        None,
        None,
        "Withholding tax",
        None,
        0,
    ])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    print(f"已导出广告费导入模板: {output_path}")


def load_original_store_lookup(original_workbook: Path | None) -> dict[str, str]:
    if original_workbook is None or not original_workbook.exists():
        return {}
    df = read_excel_cached(original_workbook, sheet_name="Tiktok订单完成表", dtype={"Order ID": str})
    if df.empty:
        return {}

    tqdm.pandas(desc="Normalize order IDs", ncols=100)
    order_ids = df["Order ID"].progress_apply(clean_id)
    tqdm.pandas(desc="Normalize store labels", ncols=100)
    store_labels = df["店铺"].progress_apply(non_empty_text)
    lookup_df = pd.DataFrame({"Order ID": order_ids, "店铺": store_labels})
    lookup_df = lookup_df[(lookup_df["Order ID"].notna()) & (lookup_df["店铺"] != "")]
    lookup_df = lookup_df.drop_duplicates(subset=["Order ID"], keep="first")
    return dict(zip(lookup_df["Order ID"], lookup_df["店铺"]))


def choose_adjustment_store(
    candidates: list[dict[str, object]],
    adjustment_id: str,
    original_store_lookup: dict[str, str],
) -> dict[str, object]:
    preferred_store = original_store_lookup.get(adjustment_id)
    if preferred_store:
        for candidate in candidates:
            if candidate["store_label"] == preferred_store:
                return candidate
    return sorted(candidates, key=lambda item: str(item["store_label"]))[0]


def build_manual_adjustment_import_df(original_workbook: Path, generated_workbook: Path) -> pd.DataFrame:
    original_df = read_excel_cached(original_workbook, sheet_name="Tiktok订单完成表", dtype={"Order ID": str}).copy()
    generated_df = read_excel_cached(generated_workbook, sheet_name="Tiktok订单完成表", dtype={"Order ID": str})
    original_df["来源行号"] = original_df.index + 2
    generated_ids = set(generated_df["Order ID"].astype(str))
    payout_ids = load_payout_id_set(generated_workbook)

    unresolved = original_df[~original_df["Order ID"].astype(str).isin(generated_ids)].copy()
    unresolved = unresolved[~unresolved["Order ID"].astype(str).isin(payout_ids)].copy()

    if unresolved.empty:
        return pd.DataFrame(columns=AD_IMPORT_HEADERS)

    direct_headers = ["Total revenue", *ORDER_EXTRA_FEE_FIELDS, "广告费", "备注", "退款原因", "二次销售(0/1)"]
    override_headers = {
        "覆盖_Total fees": "Total fees",
        "覆盖_结算金额": "结算金额",
        "覆盖_税费": "税费",
        "覆盖_总计结算金额": "总计结算金额",
        "覆盖_净利润(rmb)": "净利润(rmb)",
    }

    records = pd.DataFrame(index=unresolved.index, columns=AD_IMPORT_HEADERS)
    records["启用"] = 0
    records["记录类型"] = "新增调整行"
    tqdm.pandas(desc="Extract months", ncols=100)
    records["月份"] = unresolved["店铺"].progress_apply(lambda value: extract_month(non_empty_text(value)) or 2)
    records["店铺"] = unresolved.get("店铺")
    records["来源单据"] = "原利润表待人工补录"
    records["来源行号"] = unresolved.get("来源行号")
    records["调整_Order ID"] = unresolved.get("Order ID")
    records["调整_是否到款"] = unresolved.get("是否到款")
    records["调整_Order Status"] = unresolved.get("Order Status")
    records["调整_Seller SKU"] = unresolved.get("Seller SKU")
    quantity_series = unresolved.get("Quantity")
    records["调整_Quantity"] = quantity_series.where(quantity_series.notna(), 0) if quantity_series is not None else 0

    for header in direct_headers:
        if header not in unresolved.columns:
            continue
        source_series = unresolved[header]
        records[header] = source_series.where(source_series.notna() & (source_series != ""), None)

    for target_header, source_header in override_headers.items():
        if source_header not in unresolved.columns:
            continue
        source_series = unresolved[source_header]
        records[target_header] = source_series.where(source_series.notna() & (source_series != ""), None)

    return records.reset_index(drop=True)[AD_IMPORT_HEADERS]


def load_payout_id_set(workbook_path: Path) -> set[str]:
    payout_df = read_excel_cached(workbook_path, sheet_name="Tiktok到款明细表", dtype={"Order/adjustment ID": str})
    if "Order/adjustment ID" not in payout_df.columns:
        return set()
    payout_ids = payout_df["Order/adjustment ID"].map(clean_id)
    payout_ids = payout_ids[payout_ids.notna()]
    payout_ids = payout_ids[payout_ids.str.lower() != "order/adjustment id"]
    return set(payout_ids.tolist())


def export_prefilled_ad_import(output_path: Path, ad_import_df: pd.DataFrame) -> None:
    rows = ad_import_df.fillna("").values.tolist() if not ad_import_df.empty else []
    workbook = build_ad_import_workbook(rows)
    actual_output_path = save_workbook_with_fallback(workbook, output_path, "待人工补录调整表")
    print(f"已导出待人工补录调整表: {actual_output_path}")


def classify_misalignment_row(
    program_value: float,
    original_same_value: float,
    original_legacy_value: float,
) -> str:
    same_gap = abs(program_value - original_same_value)
    legacy_gap = abs(program_value - original_legacy_value)

    if same_gap <= MISALIGNMENT_EPSILON and legacy_gap <= MISALIGNMENT_EPSILON:
        return "两列都一致"
    if legacy_gap <= MISALIGNMENT_EPSILON and same_gap > MISALIGNMENT_EPSILON:
        return "更符合历史错位列"
    if same_gap <= MISALIGNMENT_EPSILON and legacy_gap > MISALIGNMENT_EPSILON:
        return "更符合当前真实列"
    if legacy_gap < same_gap:
        return "更接近历史错位列"
    if same_gap < legacy_gap:
        return "更接近当前真实列"
    return "两列偏差相同"


def build_manual_adjustment_observation_df(original_workbook: Path, generated_workbook: Path) -> pd.DataFrame:
    original_df = read_excel_cached(original_workbook, sheet_name="Tiktok订单完成表", dtype={"Order ID": str}).copy()
    generated_df = read_excel_cached(generated_workbook, sheet_name="Tiktok订单完成表", dtype={"Order ID": str})
    generated_ids = set(generated_df["Order ID"].astype(str))
    payout_ids = load_payout_id_set(generated_workbook)

    original_df["原表行号"] = original_df.index + 2
    tail54_ids = set(original_df.tail(54)["Order ID"].astype(str))

    unresolved = original_df[~original_df["Order ID"].astype(str).isin(generated_ids)].copy()
    unresolved = unresolved[~unresolved["Order ID"].astype(str).isin(payout_ids)].copy()

    if unresolved.empty:
        return pd.DataFrame(
            columns=[
                "人工调整层分类",
                "原表行号",
                "店铺",
                "Order ID",
                "是否到款",
                "Order Status",
                "Total revenue",
                "广告费",
                "Total fees",
                "总计结算金额",
                "备注",
            ]
        )

    unresolved["人工调整层分类"] = unresolved["Order ID"].astype(str).isin(tail54_ids).map(
        {True: "尾部54行补录", False: "散落补录"}
    )
    keep_columns = [
        "人工调整层分类",
        "原表行号",
        "店铺",
        "Order ID",
        "是否到款",
        "Order Status",
        "Total revenue",
        "广告费",
        "Total fees",
        "总计结算金额",
        "备注",
    ]
    return unresolved[keep_columns].sort_values(["人工调整层分类", "店铺", "原表行号"], kind="stable").reset_index(drop=True)


def build_manual_adjustment_summary_df(original_workbook: Path, generated_workbook: Path) -> pd.DataFrame:
    detail_df = build_manual_adjustment_observation_df(original_workbook, generated_workbook)
    if detail_df.empty:
        return pd.DataFrame(columns=["人工调整层分类", "备注", "行数", "Total revenue影响", "广告费影响", "Total fees影响", "总计结算金额影响"])

    summary = (
        detail_df.groupby(["人工调整层分类", "备注"], dropna=False)
        .agg(
            行数=("Order ID", "size"),
            Total_revenue影响=("Total revenue", "sum"),
            广告费影响=("广告费", "sum"),
            Total_fees影响=("Total fees", "sum"),
            总计结算金额影响=("总计结算金额", "sum"),
        )
        .reset_index()
        .rename(columns={"Total_revenue影响": "Total revenue影响"})
    )
    return summary.sort_values(["人工调整层分类", "行数", "备注"], ascending=[True, False, True], kind="stable").reset_index(drop=True)


def build_payout_adjustment_review_df(original_workbook: Path, generated_workbook: Path) -> pd.DataFrame:
    original_df = read_excel_cached(original_workbook, sheet_name="Tiktok订单完成表", dtype={"Order ID": str}).copy()
    generated_df = read_excel_cached(generated_workbook, sheet_name="Tiktok订单完成表", dtype={"Order ID": str})
    payout_df = read_excel_cached(generated_workbook, sheet_name="Tiktok到款明细表", dtype={"Order/adjustment ID": str})

    original_df["原表行号"] = original_df.index + 2
    generated_ids = set(generated_df["Order ID"].astype(str))
    tail54_ids = set(original_df.tail(54)["Order ID"].astype(str))
    original_order_counts = original_df["Order ID"].astype(str).value_counts()

    unresolved = original_df[~original_df["Order ID"].astype(str).isin(generated_ids)].copy()
    front20 = unresolved[~unresolved["Order ID"].astype(str).isin(tail54_ids)].copy()

    if front20.empty:
        return pd.DataFrame(
            columns=[
                "原表行号",
                "店铺",
                "Order ID",
                "原表内重复次数",
                "是否到款",
                "Order Status",
                "Total revenue",
                "广告费",
                "Total fees",
                "总计结算金额",
                "备注",
                "程序回款表是否存在",
                "回款Type",
                "Related order ID",
                "回款Total settlement amount",
                "回款Total revenue",
                "回款Total fees",
                "回款Ajustment amount",
            ]
        )

    payout_subset = payout_df.copy()
    payout_subset["Order/adjustment ID"] = payout_subset["Order/adjustment ID"].astype(str)
    payout_subset = payout_subset[payout_subset["Order/adjustment ID"].isin(front20["Order ID"].astype(str))].copy()
    payout_subset = payout_subset.rename(
        columns={
            "Order/adjustment ID": "Order ID",
            "Type": "回款Type",
            "Total settlement amount": "回款Total settlement amount",
            "Total revenue": "回款Total revenue",
            "Total fees": "回款Total fees",
            "Ajustment amount": "回款Ajustment amount",
        }
    )
    payout_subset = payout_subset[
        [
            "Order ID",
            "回款Type",
            "Related order ID",
            "回款Total settlement amount",
            "回款Total revenue",
            "回款Total fees",
            "回款Ajustment amount",
        ]
    ].drop_duplicates(subset=["Order ID"], keep="first")

    front20["原表内重复次数"] = front20["Order ID"].astype(str).map(original_order_counts)
    review_df = front20.merge(payout_subset, on="Order ID", how="left")
    review_df["程序回款表是否存在"] = review_df["回款Type"].notna()

    keep_columns = [
        "原表行号",
        "店铺",
        "Order ID",
        "原表内重复次数",
        "是否到款",
        "Order Status",
        "Total revenue",
        "广告费",
        "Total fees",
        "总计结算金额",
        "备注",
        "程序回款表是否存在",
        "回款Type",
        "Related order ID",
        "回款Total settlement amount",
        "回款Total revenue",
        "回款Total fees",
        "回款Ajustment amount",
    ]
    return review_df[keep_columns].sort_values(["店铺", "原表行号"], kind="stable").reset_index(drop=True)


def export_payout_adjustment_review_workbook(
    output_path: Path,
    original_workbook: Path,
    generated_workbook: Path,
) -> None:
    review_df = build_payout_adjustment_review_df(original_workbook, generated_workbook)
    summary_df = pd.DataFrame(
        [
            {
                "项目": "散落20条回款调整行数",
                "值": len(review_df),
                "说明": "原表独有、非尾部54行、且需要单独核对的回款调整行",
            },
            {
                "项目": "原表内重复订单数",
                "值": int((review_df["原表内重复次数"].fillna(0) > 1).sum()) if not review_df.empty else 0,
                "说明": "用于判断这20条是否可能由重复订单导致",
            },
            {
                "项目": "程序回款表可回查行数",
                "值": int(review_df["程序回款表是否存在"].fillna(False).sum()) if not review_df.empty else 0,
                "说明": "能在程序版 Tiktok到款明细表 中直接找到对应回款记录的行数",
            },
        ]
    )

    workbook = Workbook()
    workbook.remove(workbook.active)
    write_dataframe_sheet(workbook, "00-汇总", summary_df)
    write_dataframe_sheet(workbook, "前20条回款调整核对", review_df)
    actual_output_path = save_workbook_with_fallback(workbook, output_path, "前20条回款调整核对表")
    print(f"已导出前20条回款调整核对表: {actual_output_path}")


def build_misalignment_diff_df(original_workbook: Path, generated_workbook: Path) -> pd.DataFrame:
    original_df = read_excel_cached(original_workbook, sheet_name="Tiktok订单完成表", dtype={"Order ID": str})
    generated_df = read_excel_cached(generated_workbook, sheet_name="Tiktok订单完成表", dtype={"Order ID": str})
    merged = generated_df.merge(original_df, on="Order ID", how="inner", suffixes=("_程序", "_原表"))
    records: list[dict[str, object]] = []

    for real_field, legacy_field in tqdm(COLUMN_MISALIGNMENTS, desc="Analyze misaligned fee columns", unit="field", **PROGRESS_KWARGS):
        program_values = numeric_series(merged, f"{real_field}_程序")
        original_same_values = numeric_series(merged, f"{real_field}_原表")
        original_legacy_values = numeric_series(merged, f"{legacy_field}_原表")
        mask = (program_values != 0) | (original_same_values != 0) | (original_legacy_values != 0)
        subset = merged.loc[mask, ["店铺_程序", "店铺_原表", "Order ID"]].copy()
        subset["真实字段"] = real_field
        subset["原表历史落点列"] = legacy_field
        subset["程序值"] = program_values[mask].values
        subset["原表同名列值"] = original_same_values[mask].values
        subset["原表历史落点值"] = original_legacy_values[mask].values
        subset["程序-原表同名差值"] = subset["程序值"] - subset["原表同名列值"]
        subset["程序-原表历史落点差值"] = subset["程序值"] - subset["原表历史落点值"]
        same_gap = subset["程序-原表同名差值"].abs()
        legacy_gap = subset["程序-原表历史落点差值"].abs()
        subset["错位判定"] = np.select(
            [
                (same_gap <= MISALIGNMENT_EPSILON) & (legacy_gap <= MISALIGNMENT_EPSILON),
                (legacy_gap <= MISALIGNMENT_EPSILON) & (same_gap > MISALIGNMENT_EPSILON),
                (same_gap <= MISALIGNMENT_EPSILON) & (legacy_gap > MISALIGNMENT_EPSILON),
                legacy_gap < same_gap,
                same_gap < legacy_gap,
            ],
            [
                "两列都一致",
                "更符合历史错位列",
                "更符合当前真实列",
                "更接近历史错位列",
                "更接近当前真实列",
            ],
            default="两列偏差相同",
        )
        subset["历史错位改善值"] = subset["程序-原表同名差值"].abs() - subset["程序-原表历史落点差值"].abs()
        records.extend(subset.to_dict(orient="records"))

    if not records:
        return pd.DataFrame(columns=["店铺_程序", "店铺_原表", "Order ID", "真实字段", "原表历史落点列", "程序值", "原表同名列值", "原表历史落点值", "程序-原表同名差值", "程序-原表历史落点差值", "错位判定", "历史错位改善值"])

    df = pd.DataFrame(records)
    df = df.sort_values(["真实字段", "错位判定", "店铺_程序", "Order ID"], kind="stable").reset_index(drop=True)
    return df


def build_misalignment_summary_df(original_workbook: Path, generated_workbook: Path) -> pd.DataFrame:
    detail_df = build_misalignment_diff_df(original_workbook, generated_workbook)
    if detail_df.empty:
        return pd.DataFrame(columns=["真实字段", "原表历史落点列", "店铺", "错位判定", "订单数", "程序值合计", "原表同名列合计", "原表历史落点合计", "程序-原表同名差值合计", "程序-原表历史落点差值合计", "历史错位改善值合计"])

    detail_df = detail_df.copy()
    detail_df["店铺"] = detail_df["店铺_程序"].fillna("")
    empty_store_mask = detail_df["店铺"] == ""
    if empty_store_mask.any():
        detail_df.loc[empty_store_mask, "店铺"] = detail_df.loc[empty_store_mask, "店铺_原表"].fillna("")

    summary = (
        detail_df.groupby(["真实字段", "原表历史落点列", "店铺", "错位判定"], dropna=False)
        .agg(
            订单数=("Order ID", "size"),
            程序值合计=("程序值", "sum"),
            原表同名列合计=("原表同名列值", "sum"),
            原表历史落点合计=("原表历史落点值", "sum"),
            程序_原表同名差值合计=("程序-原表同名差值", "sum"),
            程序_原表历史落点差值合计=("程序-原表历史落点差值", "sum"),
            历史错位改善值合计=("历史错位改善值", "sum"),
        )
        .reset_index()
        .rename(
            columns={
                "程序_原表同名差值合计": "程序-原表同名差值合计",
                "程序_原表历史落点差值合计": "程序-原表历史落点差值合计",
            }
        )
    )
    return summary.sort_values(["真实字段", "店铺", "错位判定", "订单数"], ascending=[True, True, True, False], kind="stable").reset_index(drop=True)


def build_paid_status_diff_df(original_workbook: Path, generated_workbook: Path) -> pd.DataFrame:
    original_df = read_excel_cached(original_workbook, sheet_name="Tiktok订单完成表", dtype={"Order ID": str})
    generated_df = read_excel_cached(generated_workbook, sheet_name="Tiktok订单完成表", dtype={"Order ID": str})

    merged = original_df.merge(generated_df, on="Order ID", how="inner", suffixes=("_原表", "_程序"))
    original_paid = merged["是否到款_原表"].fillna("").astype(str)
    generated_paid = merged["是否到款_程序"].fillna("").astype(str)
    diff = merged[original_paid != generated_paid].copy()

    original_payout_ids = load_payout_id_set(original_workbook)
    generated_payout_ids = load_payout_id_set(generated_workbook)

    if diff.empty:
        return pd.DataFrame(columns=["Order ID", "店铺_原表", "店铺_程序", "是否到款_原表", "是否到款_程序", "Order Status_原表", "Order Status_程序", "备注_原表", "备注_程序", "原表回款存在", "程序回款存在"])

    diff["原表回款存在"] = diff["Order ID"].astype(str).isin(original_payout_ids)
    diff["程序回款存在"] = diff["Order ID"].astype(str).isin(generated_payout_ids)
    keep_columns = [
        "Order ID",
        "店铺_原表",
        "店铺_程序",
        "是否到款_原表",
        "是否到款_程序",
        "Order Status_原表",
        "Order Status_程序",
        "备注_原表",
        "备注_程序",
        "原表回款存在",
        "程序回款存在",
    ]
    return diff[keep_columns].sort_values(["店铺_原表", "Order ID"], kind="stable").reset_index(drop=True)


def build_pivot_group_diff_df(generated_workbook: Path, pretty_workbook: Path | None) -> pd.DataFrame:
    if pretty_workbook is None or not pretty_workbook.exists():
        return pd.DataFrame(columns=["店铺", "是否到款", "匹配情况", "Total revenue_程序", "Total revenue_精美版", "广告费_程序", "广告费_精美版", "总计结算金额_程序", "总计结算金额_精美版", "净利润(rmb)_程序", "净利润(rmb)_精美版"])

    generated_book = load_workbook(generated_workbook, read_only=True, data_only=True)
    try:
        if "透视" not in generated_book.sheetnames:
            return pd.DataFrame(columns=["店铺", "是否到款", "匹配情况", "Total revenue_程序", "Total revenue_精美版", "Total revenue_差值", "广告费_程序", "广告费_精美版", "广告费_差值", "总计结算金额_程序", "总计结算金额_精美版", "总计结算金额_差值", "净利润(rmb)_程序", "净利润(rmb)_精美版", "净利润(rmb)_差值"])
    finally:
        generated_book.close()

    generated_pivot = read_excel_cached(generated_workbook, sheet_name="透视")
    pretty_pivot = read_excel_cached(pretty_workbook, sheet_name=0)

    for df in [generated_pivot, pretty_pivot]:
        for column in ["店铺", "是否到款"]:
            if column in df.columns:
                df[column] = df[column].fillna("").astype(str)

    merged = generated_pivot.merge(pretty_pivot, on=["店铺", "是否到款"], how="outer", suffixes=("_程序", "_精美版"), indicator=True)
    merged["匹配情况"] = merged["_merge"].map({"both": "双方都有", "left_only": "仅程序版", "right_only": "仅精美版"})
    for metric in ["Total revenue", "广告费", "总计结算金额", "净利润(rmb)"]:
        merged[f"{metric}_差值"] = numeric_series(merged, f"{metric}_程序") - numeric_series(merged, f"{metric}_精美版")
    keep_columns = [
        "店铺",
        "是否到款",
        "匹配情况",
        "Total revenue_程序",
        "Total revenue_精美版",
        "Total revenue_差值",
        "广告费_程序",
        "广告费_精美版",
        "广告费_差值",
        "总计结算金额_程序",
        "总计结算金额_精美版",
        "总计结算金额_差值",
        "净利润(rmb)_程序",
        "净利润(rmb)_精美版",
        "净利润(rmb)_差值",
    ]
    return merged[keep_columns].sort_values(["匹配情况", "店铺", "是否到款"], kind="stable").reset_index(drop=True)


def write_dataframe_sheet(workbook: Workbook, sheet_name: str, dataframe: pd.DataFrame) -> None:
    worksheet = workbook.create_sheet(sheet_name)
    worksheet.freeze_panes = "A2"
    worksheet.append(list(dataframe.columns))
    for row in dataframe.itertuples(index=False, name=None):
        worksheet.append([to_excel_value(value) for value in row])

    header_fill = PatternFill(fill_type="solid", fgColor="004F81BD")
    header_font = Font(bold=True, color="00FFFFFF")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, max_col=worksheet.max_column):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top")
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"
    autosize_worksheet(worksheet, min_width=12, extra=3)


def export_comparison_workbook(
    output_path: Path,
    original_workbook: Path,
    generated_workbook: Path,
    pretty_workbook: Path | None,
    manual_adjustment_df: pd.DataFrame,
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    log_step("开始分析费用列错位明细")
    misalignment_detail_df = build_misalignment_diff_df(original_workbook, generated_workbook)
    log_step("开始汇总费用列错位影响")
    misalignment_summary_df = build_misalignment_summary_df(original_workbook, generated_workbook)
    log_step("开始提取人工调整观察明细")
    manual_observation_df = build_manual_adjustment_observation_df(original_workbook, generated_workbook)
    log_step("开始汇总人工调整影响")
    manual_summary_df = build_manual_adjustment_summary_df(original_workbook, generated_workbook)
    log_step("开始分析是否到款差异")
    paid_status_diff_df = build_paid_status_diff_df(original_workbook, generated_workbook)
    log_step("开始分析透视分组差异")
    pivot_group_diff_df = build_pivot_group_diff_df(generated_workbook, pretty_workbook)

    summary_df = pd.DataFrame(
        [
            {"项目": "第1项-人工调整观察行数", "值": len(manual_observation_df), "说明": "原表独有且程序当前未恢复的调整层，已拆成尾部54行补录和散落补录"},
            {"项目": "第2项-人工调整分类数", "值": len(manual_summary_df), "说明": "按 人工调整层分类 + 备注 汇总后的影响分组"},
            {"项目": "第3项-费用列错位差异行数", "值": len(misalignment_detail_df), "说明": "按 4 组历史错位列映射展开后的逐行差异"},
            {"项目": "第3项-费用列错位汇总组数", "值": len(misalignment_summary_df), "说明": "按 真实字段 + 历史落点列 + 错位判定 汇总"},
            {"项目": "第4项-是否到款差异订单数", "值": len(paid_status_diff_df), "说明": "同一 Order ID 在原表与程序版的是否到款不同"},
            {"项目": "第2项-待人工补录行数", "值": len(manual_adjustment_df), "说明": "程序当前无法从系统回款直接恢复，已预填到待人工补录表"},
            {"项目": "精美版分组差异行数", "值": len(pivot_group_diff_df), "说明": "程序透视与精美版按 店铺+是否到款 的分组差异"},
        ]
    )
    write_dataframe_sheet(workbook, "00-汇总", summary_df)
    write_dataframe_sheet(workbook, "1-人工调整观察", manual_observation_df)
    write_dataframe_sheet(workbook, "1-人工调整汇总", manual_summary_df)
    write_dataframe_sheet(workbook, "3-费用列错位差异", misalignment_detail_df)
    write_dataframe_sheet(workbook, "3-费用列错位汇总", misalignment_summary_df)
    write_dataframe_sheet(workbook, "4-是否到款差异", paid_status_diff_df)
    write_dataframe_sheet(workbook, "4-透视分组差异", pivot_group_diff_df)

    actual_output_path = save_workbook_with_fallback(workbook, output_path, "3和4差异核对表")
    print(f"已导出3和4差异核对表: {actual_output_path}")


def build_analysis_dataframe(order_workbook: Path, exchange_rate: float) -> pd.DataFrame:
    df = read_excel_cached(order_workbook, sheet_name="Tiktok订单完成表")
    group_keys = ["店铺", "是否到款"]
    prepared = df[group_keys].copy()
    agg_map: dict[str, str | callable] = {}

    for header in ANALYSIS_HEADERS[2:]:
        if header in {"总收入(RMB)", "利润率"}:
            continue

        source_column = resolve_analysis_source_column(df, header)
        if header in ANALYSIS_FIRST_VALUE_HEADERS:
            prepared[header] = df[source_column] if source_column else None
            agg_map[header] = first_non_empty
        else:
            prepared[header] = pd.to_numeric(df[source_column], errors="coerce").fillna(0) if source_column else 0.0
            agg_map[header] = "sum"

    grouped = prepared.groupby(group_keys, dropna=False, sort=True).agg(agg_map).reset_index()
    grouped["总收入(RMB)"] = grouped["Total revenue"] / exchange_rate
    grouped["利润率"] = np.where(
        grouped["总收入(RMB)"].fillna(0) != 0,
        grouped["净利润(rmb)"].fillna(0) / grouped["总收入(RMB)"],
        0.0,
    )
    return grouped[ANALYSIS_HEADERS]


def populate_analysis_sheet(worksheet, analysis_df: pd.DataFrame) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="004F81BD")
    header_font = Font(bold=True, color="00FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    data_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    worksheet.title = "透视"
    worksheet.freeze_panes = "C2"
    worksheet.append(ANALYSIS_HEADERS)

    numeric_columns = set(range(3, 50))
    text_columns = {50, 51, 52}
    for row_values in analysis_df.itertuples(index=False, name=None):
        worksheet.append([to_excel_value(value) for value in row_values])

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for column_index in range(1, len(ANALYSIS_HEADERS) + 1):
        worksheet.column_dimensions[get_column_letter(column_index)].width = 20

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, max_col=worksheet.max_column):
        for cell in row:
            cell.border = data_border
            if cell.column == 48:
                cell.number_format = "0.00%"
            elif cell.column in numeric_columns:
                cell.number_format = "#,##0.00"
            elif cell.column in text_columns:
                cell.number_format = "@"

def write_analysis_sheet(output_path: Path, analysis_df: pd.DataFrame) -> Path:
    workbook = load_workbook(output_path)
    if "透视" in workbook.sheetnames:
        worksheet = workbook["透视"]
        workbook.remove(worksheet)
    worksheet = workbook.create_sheet("透视", 0)
    populate_analysis_sheet(worksheet, analysis_df)
    actual_output_path = save_workbook_with_fallback(workbook, output_path, "静态透视工作表")
    if actual_output_path != output_path:
        print(f"警告: 透视页已写入备用文件: {actual_output_path}")
    return actual_output_path


def add_analysis_sheet_only(output_path: Path, exchange_rate: float) -> Path:
    analysis_df = build_analysis_dataframe(output_path, exchange_rate)
    actual_output_path = write_analysis_sheet(output_path, analysis_df)
    print(f"已更新静态“透视”工作表: {actual_output_path}")
    return actual_output_path


def load_cost_map(cost_workbook: Path) -> dict[str, float]:
    wb = load_workbook(cost_workbook, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    costs: dict[str, float] = {}
    for sku, total_cost in ws.iter_rows(min_row=2, values_only=True):
        key = clean_id(sku)
        number = to_number(total_cost)
        if key and number is not None:
            costs[key] = number
    return costs


def extract_month(text: str) -> int | None:
    match = re.search(r"(\d{1,2})月", text)
    if not match:
        return None
    return int(match.group(1))


def extract_category(text: str) -> str:
    category = text
    category = category.replace("菲律宾 TK ", "")
    category = category.replace("菲律宾TK ", "")
    category = category.replace("菲律宾 ", "")
    category = re.sub(r"\d{4}年\d{1,2}月(?:到\d{4}年\d{1,2}月\d{1,2}日|到\d{1,2}月\d{1,2}日)?", "", category)
    category = category.replace("店铺数据", "")
    category = category.replace("店铺", "")
    category = category.replace("订单", "")
    category = category.replace("回款", "")
    category = category.replace("数据", "")
    category = category.replace(" ", "")
    return category.strip("-_")


def build_store_label(path: Path, system_dir: Path) -> str:
    relative_parts = path.relative_to(system_dir).parts
    parent_parts = relative_parts[:-1]
    month = extract_month(path.name) or next((extract_month(part) for part in reversed(parent_parts) if extract_month(part)), None) or 2

    if parent_parts and "-菲律宾TK" in parent_parts[0]:
        owner = parent_parts[0].split("-", 1)[0].strip()
        category_source = parent_parts[1] if len(parent_parts) > 1 else path.stem
    else:
        category_source = parent_parts[0] if parent_parts else path.stem
        category_guess = extract_category(category_source)
        owner = TOP_LEVEL_OWNER_MAP.get(category_guess)
        if owner is None:
            raise ValueError(f"无法从路径推断负责人: {path}")

    category = extract_category(category_source)
    return f"{month}月份-{owner}{category}"


def discover_source_files(system_dir: Path) -> list[SourceFile]:
    files: list[SourceFile] = []
    all_paths = list(system_dir.rglob("*.xlsx"))
    for path in tqdm(all_paths, desc="Discover source files", unit="file", **PROGRESS_KWARGS):
        if path.name.startswith("~$"):
            continue
        name = path.name
        if "订单" in name:
            kind = "order"
        elif "回款" in name:
            kind = "payout"
        elif "营销费+退款统计" in name:
            kind = "marketing"
        else:
            continue
        files.append(SourceFile(path=path, kind=kind, store_label=build_store_label(path, system_dir)))
    return sorted(files, key=lambda item: (item.kind, item.store_label, str(item.path)))


def open_worksheet(
    path: Path,
    sheet_name: str | None = None,
    required_headers: Iterable[str] = (),
):
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    worksheet.reset_dimensions()
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    header_keys = {normalize_header(value) for value in header_row if value is not None}
    required_keys = {normalize_header(value) for value in required_headers}

    if required_keys and not required_keys.issubset(header_keys):
        workbook = load_workbook(path, read_only=False, data_only=True)
        worksheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))

    return worksheet, header_row


def build_payout_summary(
    payout_files: list[SourceFile],
    payout_sheet,
    valid_order_ids: set[str],
    original_store_lookup: dict[str, str],
) -> tuple[dict[str, dict[str, float | str | None]], int, list[StoreAdjustment]]:
    payout_sheet.append(clean_row(PAYOUT_HEADERS))
    payout_index = {normalize_header(header): idx for idx, header in enumerate(PAYOUT_HEADERS[1:])}
    aggregates: dict[str, dict[str, float | str | None]] = defaultdict(dict)
    adjustments: list[StoreAdjustment] = []
    adjustment_candidates: dict[tuple[str, str, str], dict[str, object]] = {}
    row_count = 0

    for source in tqdm(payout_files, desc="Summarize payout files", unit="file", **PROGRESS_KWARGS):
        ws, header_row = open_worksheet(
            source.path,
            "Order details",
            required_headers=[
                "Order/adjustment ID",
                "Type",
                "Affiliate commission deposit",
                "Campaign resource fee",
                "Order processing fee",
                "Related order ID",
            ],
        )
        source_index = {normalize_header(header): idx for idx, header in enumerate(header_row)}

        for row in ws.iter_rows(min_row=2, values_only=True):
            combined_row = [source.store_label] + [None] * (len(PAYOUT_HEADERS) - 1)
            row_values: dict[str, object] = {}

            for header in PAYOUT_HEADERS[1:]:
                normalized = normalize_header(header)
                source_idx = source_index.get(normalized)
                value = row[source_idx] if source_idx is not None and source_idx < len(row) else None
                combined_row[payout_index[normalized] + 1] = value
                row_values[header] = value

            payout_sheet.append(clean_row(combined_row))
            row_count += 1

            payout_type = str(row_values.get("Type") or "").strip()
            rule = PAYOUT_ADJUSTMENT_RULES.get(payout_type)
            if rule:
                total_settlement_amount = to_number(row_values.get("Total settlement amount"))
                adjustment_id = clean_id(row_values.get("Order/adjustment ID"))
                source_month = extract_month(source.path.name) or 2
                if adjustment_id and total_settlement_amount not in (None, 0):
                    if rule["match_source_month"] and not matches_source_month(row_values, source_month):
                        continue
                    bucket_key = (payout_type, source.store_label, adjustment_id)
                    bucket = adjustment_candidates.setdefault(
                        bucket_key,
                        {
                            "store_label": source.store_label,
                            "month": source_month,
                            "order_id": adjustment_id,
                            "amount": 0.0,
                            "note": rule["note"],
                            "source_file": source.path.name,
                            "source_type": rule["source_type"],
                            "target_header": rule["target_header"],
                            "paid_status": rule["paid_status"],
                        },
                    )
                    bucket["amount"] = float(bucket["amount"]) + float(total_settlement_amount)
                continue

            order_key = resolve_payout_order_key(row_values, valid_order_ids)
            if not order_key:
                continue

            bucket = aggregates[order_key]
            bucket["店铺"] = source.store_label
            for order_field, payout_field in PAYOUT_FIELD_ALIASES.items():
                number = to_number(row_values.get(payout_field))
                if number is None:
                    continue
                bucket[order_field] = to_number(bucket.get(order_field)) or 0.0
                bucket[order_field] = float(bucket[order_field]) + number

    grouped_candidates: dict[tuple[str, str], list[dict[str, object]]] = defaultdict(list)
    for (payout_type, _store_label, adjustment_id), candidate in adjustment_candidates.items():
        grouped_candidates[(payout_type, adjustment_id)].append(candidate)

    for (_payout_type, adjustment_id), candidates in grouped_candidates.items():
        chosen = choose_adjustment_store(candidates, adjustment_id, original_store_lookup)
        adjustments.append(
            StoreAdjustment(
                store_label=str(chosen["store_label"]),
                month=int(chosen["month"]),
                order_id=str(chosen["order_id"]),
                amount=float(chosen["amount"]),
                note=str(chosen["note"]),
                source_file=str(chosen["source_file"]),
                source_type=str(chosen["source_type"]),
                target_header=str(chosen["target_header"]),
                paid_status=str(chosen["paid_status"]),
            )
        )

    return aggregates, row_count, adjustments


def build_order_summary(
    order_files: list[SourceFile],
    payout_aggregates: dict[str, dict[str, float | str | None]],
    cost_map: dict[str, float],
    tax_rate: float,
    exchange_rate: float,
    include_canceled: bool,
) -> list[list[object]]:
    order_rows: list[list[object]] = []

    for source in tqdm(order_files, desc="Summarize order files", unit="file", **PROGRESS_KWARGS):
        ws, header_row = open_worksheet(source.path, required_headers=["Order ID", "Seller SKU", "Quantity"])
        source_index = {normalize_header(header): idx for idx, header in enumerate(header_row)}

        for row in ws.iter_rows(min_row=3, values_only=True):
            order_id = clean_id(row[source_index[normalize_header("Order ID")]])
            if not order_id:
                continue

            order_status = row[source_index[normalize_header("Order Status")]]
            if not include_canceled and order_status in {"Canceled", "Cancelled"}:
                continue

            seller_sku = clean_id(row[source_index[normalize_header("Seller SKU")]])
            quantity = to_number(row[source_index[normalize_header("Quantity")]]) or 0.0
            payout = payout_aggregates.get(order_id, {})

            total_revenue = payout.get("Total revenue")
            if total_revenue is None:
                total_revenue = safe_sum([
                    row[source_index[normalize_header("SKU Subtotal After Discount")]],
                    -1 * (to_number(row[source_index[normalize_header("Shipping Fee Seller Discount")]]) or 0.0),
                    -1 * (to_number(row[source_index[normalize_header("Taxes")]]) or 0.0),
                ])

            sku_cost = cost_map.get(seller_sku) if seller_sku else None
            total_cost = sku_cost * quantity if sku_cost is not None else None

            fee_values = [payout.get(field) for field in ORDER_EXTRA_FEE_FIELDS]
            ad_fee = None
            total_fees = safe_sum([*fee_values, ad_fee])
            settled_amount = safe_sum([total_revenue, total_fees])
            tax_fee = settled_amount * tax_rate if settled_amount is not None else None
            final_settlement = settled_amount - tax_fee if settled_amount is not None else None
            net_profit = (final_settlement / exchange_rate - total_cost) if final_settlement is not None and total_cost is not None else None

            source_values = list(row)
            combined_row = [
                source.store_label,
                order_id,
                order_id,
                "是" if order_id in payout_aggregates else "否",
                *source_values[1:55],
                total_revenue,
                sku_cost,
                total_cost,
                *fee_values,
                ad_fee,
                total_fees,
                settled_amount,
                tax_fee,
                final_settlement,
                net_profit,
                source_values[57] if len(source_values) > 57 else None,
                None,
                source_values[55] if len(source_values) > 55 else None,
                source_values[56] if len(source_values) > 56 else None,
            ]
            order_rows.append(clean_row(combined_row))

    return order_rows


def build_workbook(
    system_dir: Path,
    cost_workbook: Path,
    ad_import_path: Path | None,
    output_path: Path,
    original_workbook: Path | None,
    pretty_workbook: Path | None,
    manual_adjustment_output: Path | None,
    comparison_output: Path | None,
    payout_adjustment_review_output: Path | None,
    tax_rate: float,
    exchange_rate: float,
    include_canceled: bool,
    add_analysis_sheet: bool,
) -> None:
    log_step("开始扫描源文件")
    sources = discover_source_files(system_dir)
    order_files = [item for item in sources if item.kind == "order"]
    payout_files = [item for item in sources if item.kind == "payout"]
    marketing_files = [item for item in sources if item.kind == "marketing"]

    if not order_files or not payout_files:
        raise ValueError("未找到完整的订单表和回款表。")

    log_step("开始读取成本表")
    cost_map = load_cost_map(cost_workbook)
    log_step("开始收集有效订单 ID")
    valid_order_ids = collect_order_ids(order_files, include_canceled)
    log_step("开始建立原表店铺归属映射")
    original_store_lookup = load_original_store_lookup(original_workbook)

    workbook = Workbook(write_only=True)
    order_sheet = workbook.create_sheet("Tiktok订单完成表")
    payout_sheet = workbook.create_sheet("Tiktok到款明细表")

    log_step("开始汇总回款文件")
    payout_aggregates, payout_rows, payout_adjustments = build_payout_summary(
        payout_files,
        payout_sheet,
        valid_order_ids,
        original_store_lookup,
    )
    log_step("开始汇总订单文件")
    order_rows = build_order_summary(
        order_files,
        payout_aggregates,
        cost_map,
        tax_rate,
        exchange_rate,
        include_canceled,
    )
    log_step("开始提取营销调整")
    marketing_adjustments = build_marketing_store_adjustments(marketing_files)
    order_rows, system_adjustment_stats = apply_system_adjustments(
        order_rows,
        [*payout_adjustments, *marketing_adjustments],
        tax_rate,
        exchange_rate,
    )
    log_step("开始读取广告费导入表")
    ad_import_df = load_ad_import(ad_import_path)
    order_rows, ad_import_stats = apply_ad_import(order_rows, ad_import_df, tax_rate, exchange_rate)

    order_sheet.append(clean_row(ORDER_HEADERS))
    for row in order_rows:
        order_sheet.append(row)

    log_step("开始写出主工作簿")
    actual_output_path = save_workbook_with_fallback(workbook, output_path, "程序生成工作簿")
    workbook.close()
    gc.collect()

    if add_analysis_sheet:
        log_step("开始生成静态透视")
        actual_output_path = add_analysis_sheet_only(actual_output_path, exchange_rate)

    manual_adjustment_df = pd.DataFrame(columns=AD_IMPORT_HEADERS)
    if original_workbook is not None and original_workbook.exists() and manual_adjustment_output is not None:
        log_step("开始生成待人工补录调整表")
        manual_adjustment_df = build_manual_adjustment_import_df(original_workbook, actual_output_path)
        export_prefilled_ad_import(manual_adjustment_output, manual_adjustment_df)

    if original_workbook is not None and original_workbook.exists() and comparison_output is not None:
        log_step("开始生成差异核对表")
        export_comparison_workbook(comparison_output, original_workbook, actual_output_path, pretty_workbook, manual_adjustment_df)

    if original_workbook is not None and original_workbook.exists() and payout_adjustment_review_output is not None:
        log_step("开始生成前20条回款调整核对表")
        export_payout_adjustment_review_workbook(payout_adjustment_review_output, original_workbook, actual_output_path)

    print(f"已生成: {actual_output_path}")
    print(f"订单文件: {len(order_files)} 个, 汇总行数: {len(order_rows)}")
    print(f"回款文件: {len(payout_files)} 个, 汇总行数: {payout_rows}")
    print(f"营销统计文件: {len(marketing_files)} 个")
    print(
        "系统调整行: "
        f"线上广告费 {system_adjustment_stats.get('online_ad', 0)} 条, "
        f"线下广告费 {system_adjustment_stats.get('offline_ad', 0)} 条, "
        f"线下退款 {system_adjustment_stats.get('offline_refund', 0)} 条, "
        f"Withholding tax {system_adjustment_stats.get('withholding_tax', 0)} 条, "
        f"Platform reimbursement {system_adjustment_stats.get('platform_reimbursement', 0)} 条, "
        f"Logistics reimbursement {system_adjustment_stats.get('logistics_reimbursement', 0)} 条, "
        f"Violation fee {system_adjustment_stats.get('violation_fee', 0)} 条"
    )
    print(
        "广告费导入规则: "
        f"分摊 {ad_import_stats['allocation_rules']} 条, "
        f"新增调整行 {ad_import_stats['adjustment_rows']} 条"
    )
    if add_analysis_sheet:
        print("已生成静态“透视”工作表，样式对齐到盈利分析详细精美版模板。")
    if manual_adjustment_output is not None and original_workbook is not None and original_workbook.exists():
        print(f"已输出待人工补录调整表: {manual_adjustment_output}")
    if comparison_output is not None and original_workbook is not None and original_workbook.exists():
        print(f"已输出3和4差异核对表: {comparison_output}")
    if payout_adjustment_review_output is not None and original_workbook is not None and original_workbook.exists():
        print(f"已输出前20条回款调整核对表: {payout_adjustment_review_output}")
    print("说明: 当前脚本已自动处理订单、回款、SKU 成本与可选广告费导入表。默认跳过取消订单。")


def parse_args() -> argparse.Namespace:
    workspace = Path(__file__).resolve().parents[1]
    generated_dir = build_generated_dir(workspace)
    parser = argparse.ArgumentParser(description="整合 TikTok 订单表和回款表，生成汇总工作簿。")
    parser.add_argument(
        "--system-dir",
        type=Path,
        default=workspace / "01-TK文件" / "tk" / "011-系统数据",
        help="011-系统数据 文件夹路径",
    )
    parser.add_argument(
        "--cost-workbook",
        type=Path,
        default=workspace / "成本表.xlsx",
        help="SKU 成本表路径",
    )
    parser.add_argument(
        "--ad-import",
        type=Path,
        default=generated_dir / "广告费导入表.xlsx",
        help="广告费导入表路径；不存在时按无补充表处理",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=generated_dir / "2026年菲律宾2月tiktok订单利润表_程序生成.xlsx",
        help="输出工作簿路径，默认写入 01-TK文件/程序生成结果",
    )
    parser.add_argument(
        "--original-workbook",
        type=Path,
        default=workspace / "01-TK文件" / "2026年菲律宾2月tiktok订单利润表（1.1%）.xlsx",
        help="原利润表路径；存在时用于补重复调整归属、导出人工补录表和差异核对表",
    )
    parser.add_argument(
        "--pretty-workbook",
        type=Path,
        default=workspace.parent / "盈利分析详细精美版.xlsx",
        help="精美版样板路径；存在时导出透视分组差异",
    )
    parser.add_argument(
        "--manual-adjustment-output",
        type=Path,
        default=generated_dir / "待人工补录调整.xlsx",
        help="待人工补录调整表输出路径",
    )
    parser.add_argument(
        "--comparison-output",
        type=Path,
        default=generated_dir / "3和4差异核对.xlsx",
        help="第3项和第4项差异核对表输出路径",
    )
    parser.add_argument(
        "--payout-adjustment-review-output",
        type=Path,
        default=generated_dir / "前20条回款调整核对.xlsx",
        help="散落20条回款调整核对表输出路径",
    )
    parser.add_argument("--tax-rate", type=float, default=0.011, help="税率，默认 1.1%%")
    parser.add_argument("--exchange-rate", type=float, default=8.4672, help="PHP 对 RMB 汇率")
    parser.add_argument("--include-canceled", action="store_true", help="保留取消订单")
    parser.add_argument("--skip-analysis-sheet", action="store_true", help="不生成静态透视工作表")
    parser.add_argument("--analysis-only", action="store_true", help="仅基于现有输出文件生成静态透视工作表")
    parser.add_argument("--export-ad-template", action="store_true", help="导出广告费导入表模板后退出")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if args.export_ad_template:
        export_ad_import_template(args.ad_import)
        return

    if args.analysis_only:
        add_analysis_sheet_only(args.output, args.exchange_rate)
        return

    build_workbook(
        system_dir=args.system_dir,
        cost_workbook=args.cost_workbook,
        ad_import_path=args.ad_import,
        output_path=args.output,
        original_workbook=args.original_workbook,
        pretty_workbook=args.pretty_workbook,
        manual_adjustment_output=args.manual_adjustment_output,
        comparison_output=args.comparison_output,
        payout_adjustment_review_output=args.payout_adjustment_review_output,
        tax_rate=args.tax_rate,
        exchange_rate=args.exchange_rate,
        include_canceled=args.include_canceled,
        add_analysis_sheet=not args.skip_analysis_sheet,
    )


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n已手动中断执行（KeyboardInterrupt）。脚本本身未发生逻辑异常，可稍后重新运行。")