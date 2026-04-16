import openpyxl
import os

file_path = r'C:\Users\Bin\Desktop\东海表格\2026年菲律宾2月tiktok订单利润表（1.1%）.xlsx'
if not os.path.exists(file_path):
    print(f'Error: File not found at {file_path}')
    exit(1)

# 加载工作簿
try:
    wb = openpyxl.load_workbook(file_path, data_only=False)
except Exception as e:
    print(f'Error loading workbook: {e}')
    exit(1)

sheet_name = '透视'
if sheet_name not in wb.sheetnames:
    print(f'Error: Worksheet "{sheet_name}" not found. Available sheets: {wb.sheetnames}')
    exit(1)

ws = wb[sheet_name]

# 1. 检查全局外部链接 (External Links)
print('--- External Links ---')
external_links = []
# openpyxl 的外部链接存储在 _external_links
if hasattr(wb, '_external_links'):
    for link in wb._external_links:
        print(f'Found global external link: {link.file_link.Target}')
        external_links.append(link.file_link.Target)
else:
    print('No global external links found via standard API.')

# 2. 检查单元格公式中的外部引用
print('\n--- Formulas with External References ---')
external_formulas_count = 0
for row in ws.iter_rows():
    for cell in row:
        val = str(cell.value) if cell.value else ""
        # 检查公式标志 'f' 且包含外部工作簿引用格式 '[' 或绝对路径 ':\\'
        if cell.data_type == 'f' and ('[' in val or ':\\' in val or '!' in val):
            # 排除本文件内的引用（本文件引用通常是 SheetName!A1）
            # 外部引用通常是 [ExternalFile.xlsx]Sheet!A1
            if '[' in val:
                print(f'Cell {cell.coordinate} has external formula: {val}')
                external_formulas_count += 1

if external_formulas_count == 0:
    print('No external cell formulas detected in "透视" sheet.')

# 3. 检查数据透视表 (Pivot Tables)
print('\n--- Pivot Table Sources ---')
pivot_tables = ws._pivots
if pivot_tables:
    print(f'Found {len(pivot_tables)} pivot table(s) on "透视" sheet.')
    for i, pt in enumerate(pivot_tables):
        print(f'Pivot Table [{i+1}]: {pt.name}')
        # 数据透视表的数据源通常在 CacheDefinition 中
        # openpyxl 访问 cache 的方式
        try:
            cache = pt.cache
            # 如果是 WorksheetSource，通常有 ref 和 sheet
            if cache.worksheetsource:
                print(f'  Source: Sheet "{cache.worksheetsource.sheet}", Ref: "{cache.worksheetsource.ref}"')
            else:
                print('  Source: Not a simple worksheet source (possibly external or power query).')
        except Exception as e:
            print(f'  Could not extract detailed source: {e}')
else:
    print('No active Pivot Tables found on the "透视" sheet (it may contain static data).')

# 4. 检查是否有定义名称 (Defined Names) 指向外部
print('\n--- Defined Names ---')
for name in wb.defined_names.definedName:
    if '[' in str(name.value):
        print(f'Defined name "{name.name}" points to: {name.value}')
