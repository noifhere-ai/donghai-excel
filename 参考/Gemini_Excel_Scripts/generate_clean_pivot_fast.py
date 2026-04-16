import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from collections import defaultdict

file_path = r'C:\Users\Bin\Desktop\东海表格\2026年菲律宾2月tiktok订单利润表（1.1%）.xlsx'

# 1. 定义费用归类
platform_fee_cols = ['Transaction fee', 'TikTok Shop commission fee', 'Order processing fee', 'Affiliate commission deposit']
logistics_fee_cols = ['Seller shipping fee', 'Shipping Service Fee']
marketing_fee_cols = [
    'Affiliate commission', 'Affiliate Shop Ads commission\t', 'Bonus cashback service fee', 
    'LIVE Specials service fee', 'Campaign resource fee', 'EAMS Program service fee', 
    'GMV Max Coupon', '广告费'
]
base_cols = ['Total revenue', '总计结算金额', 'SKU总成本(rmb)', '净利润(rmb)', '税费']

def fast_process():
    print("Opening workbook (readonly mode)...")
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb['Tiktok订单完成表']
    
    # 获取表头索引
    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    col_map = {name: i for i, name in enumerate(headers) if name}
    
    data_map = defaultdict(lambda: defaultdict(float))
    
    # 逐行读取核心列
    print("Reading rows...")
    count = 0
    for row in ws.iter_rows(min_row=2):
        store = row[col_map['店铺']].value
        status = row[col_map['是否到款']].value
        key = (store, status)
        
        # 累加各项指标
        for name in base_cols + platform_fee_cols + logistics_fee_cols + marketing_fee_cols:
            if name in col_map:
                val = row[col_map[name]].value
                try:
                    data_map[key][name] += float(val) if val else 0
                except: pass
        
        count += 1
        if count % 5000 == 0:
            print(f"Processed {count} rows...")
            
    wb.close()
    
    # 2. 计算汇总结果
    results = []
    for (store, status), vals in data_map.items():
        if not store: continue
        
        platform = sum(vals[c] for c in platform_fee_cols)
        logistics = sum(vals[c] for c in logistics_fee_cols)
        marketing = sum(vals[c] for c in marketing_fee_cols)
        
        # 利润率计算 (估算汇率 7.8)
        revenue = vals['Total revenue']
        profit = vals['净利润(rmb)']
        margin = profit / (revenue / 7.8) if revenue != 0 else 0
        
        results.append([
            store, status, revenue, vals['总计结算金额'], 
            platform, logistics, marketing, vals['税费'],
            vals['SKU总成本(rmb)'], profit, margin
        ])
    
    # 3. 写入新表
    print("Writing results to new sheet...")
    wb_write = openpyxl.load_workbook(file_path)
    if '盈利分析精简版' in wb_write.sheetnames:
        del wb_write['盈利分析精简版']
    new_ws = wb_write.create_sheet('盈利分析精简版')
    
    headers_final = ['店铺名称', '结算状态', '总收入(PHP)', '到账金额(PHP)', '平台费', '物流费', '营销费', '税费', '成本(RMB)', '净利润(RMB)', '利润率']
    new_ws.append(headers_final)
    
    # 样式
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for r_idx, row_data in enumerate(results, start=2):
        new_ws.append(row_data)
        
    # 美化格式
    for col in new_ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            cell.border = border
            cell.alignment = center_align
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = bold_font
            else:
                if column_letter in 'CDEFGHIJ':
                    cell.number_format = '#,##0.00'
                elif column_letter == 'K':
                    cell.number_format = '0.00%'
            
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        new_ws.column_dimensions[column_letter].width = max_length + 2

    wb_write.save(file_path)
    print("Done!")

if __name__ == "__main__":
    fast_process()
